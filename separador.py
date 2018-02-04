from openpyxl import Workbook
from openpyxl import load_workbook

def abrirPlanilha(nomeArquivo):
    planilha = load_workbook(nomeArquivo)
    planilha = planilha.active()
    return tuple(planilha.rows)

def separarLinhas(matrizPlanilha, colunaVer):
    categorias = []
    competidores = []
    for linha in matrizPlanilha:
        if not linha[colunaVer] in categorias:
            categorias.append(linha[colunaVer])
            competidores.append([])
        competidores[categorias.index(linha[colunaVer])].append(linha)
    linha = 1
    for categoria in competidores:
        planilhaNova = Workbook()
        planilhaNovaAtiva = planilhaNova.active()
        for competidor in categoria:
            coluna = 1
            for info in competidor:
                planilhaNovaAtiva.cell(row=linha, column=coluna).value = info
                coluna += 1
        nomeArq = categorias[competidores.index(categoria)] + ".xlsx"
        print(nomeArq+"criado!")
        planilhaNova.save(nomeArq)
        linha += 1